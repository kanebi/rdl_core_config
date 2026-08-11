# -*- coding: utf-8 -*-
"""Shared helpers for RDL_Trading_Odoo.xlsx template import."""
from __future__ import annotations

import io
import logging
import re
from typing import Any

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Sheet indexes in RDL_Trading_Odoo.xlsx (0-based)
SHEET_PRODUCT_MASTER = 1
SHEET_OPENING_INVENTORY = 3
DEFAULT_HEADER_ROW_1INDEXED = 4


def require_openpyxl():
    if openpyxl is None:
        from odoo.exceptions import UserError
        raise UserError("Please install the 'openpyxl' Python library to import Excel files.")


def load_workbook(file_bytes: bytes):
    require_openpyxl()
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)


def normalize_header(value) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    text = text.replace('₦', '').replace('\u20a6', '')
    text = re.sub(r'\(\s*\)', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def normalize_key(value) -> str:
    return normalize_header(value).lower()


def safe_float(value, default=0.0) -> float:
    if value is None:
        return default
    text = str(value).strip()
    if not text or text.lower() == 'nan':
        return default
    try:
        return float(text.replace(',', ''))
    except (TypeError, ValueError):
        return default


def safe_str(value, default='') -> str:
    if value is None:
        return default
    text = str(value).strip()
    if not text or text.lower() == 'nan':
        return default
    return text


def resolve_sheet(workbook, *, sheet_index=None, sheet_name=None):
    if sheet_name:
        name = sheet_name.strip()
        if name.isdigit():
            sheet_index = int(name)
        elif name in workbook.sheetnames:
            return workbook[name]
        else:
            from odoo.exceptions import UserError
            raise UserError(("Sheet '%s' not found in workbook.") % name)
    if sheet_index is None:
        sheet_index = 0
    if sheet_index < 0 or sheet_index >= len(workbook.worksheets):
        from odoo.exceptions import UserError
        raise UserError(("Sheet index %s is out of bounds.") % sheet_index)
    return workbook.worksheets[sheet_index]


def iter_data_rows(sheet, header_row_1indexed=DEFAULT_HEADER_ROW_1INDEXED):
    """Yield dict rows keyed by normalized header labels."""
    rows = list(sheet.iter_rows(values_only=True))
    header_idx = header_row_1indexed - 1
    if header_idx < 0 or header_idx >= len(rows):
        from odoo.exceptions import UserError
        raise UserError(("Header row %s is out of bounds.") % header_row_1indexed)

    raw_headers = rows[header_idx]
    headers = [normalize_header(h) for h in raw_headers]
    header_keys = [normalize_key(h) for h in headers]

    for row_idx in range(header_idx + 1, len(rows)):
        row_values = rows[row_idx]
        row = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = row_values[col_idx] if col_idx < len(row_values) else None
            row[header_keys[col_idx]] = row[header]
        yield row_idx + 1, row


def pick(row: dict, *labels, default=None):
    """Get a cell value using exact or normalized header labels."""
    for label in labels:
        if label in row:
            val = row[label]
            if val is not None and str(val).strip() not in ('', 'nan'):
                return val
        key = normalize_key(label)
        if key in row:
            val = row[key]
            if val is not None and str(val).strip() not in ('', 'nan'):
                return val
    return default


def pick_float(row: dict, *labels, default=0.0) -> float:
    return safe_float(pick(row, *labels, default=None), default=default)


def infer_pack_uom_type(uom_label: str, unit_packaging: str) -> str:
    uom = (uom_label or '').lower()
    packaging = (unit_packaging or '').lower()
    if 'carton' in uom or packaging in ('can', 'plastic', 'pet'):
        return 'carton'
    if 'case' in uom:
        return 'case'
    return 'crate'


def build_product_vals(row: dict) -> dict | None:
    """
    Build product.template values from a Product Master row.
    Uses whole Sales Price / Cost Price only (no empties/BOM split fields).
    """
    sku = safe_str(pick(row, 'SKU'))
    if not sku:
        return None

    name = safe_str(pick(row, 'Product Name'), sku)
    unit_packaging = safe_str(pick(row, 'Unit Packaging')).lower()
    uom_label = safe_str(pick(row, 'UOM'))
    categ_name = safe_str(pick(row, 'Category'))

    list_price = pick_float(row, 'Sales Price (₦)', 'Sales Price')
    standard_price = pick_float(row, 'Cost Price (₦)', 'Cost Price')
    if not standard_price:
        standard_price = pick_float(row, 'Full Crate Cost Price')

    vals = {
        'name': name,
        'default_code': sku,
        'type': 'consu',
        'is_storable': True,
        'available_in_pos': True,
        'sale_ok': True,
        'purchase_ok': True,
        'list_price': list_price,
        'standard_price': standard_price,
    }

    pack_qty = pick_float(row, 'Bottles in a Crate', 'Units per Pack', 'Pack Qty')
    uses_pack = unit_packaging in ('bottle', 'can', 'plastic', 'pet', 'carton') or bool(pack_qty)

    if uses_pack:
        if not pack_qty:
            pack_qty = 24.0
        vals.update({
            'pack_qty': pack_qty,
            'pack_uom_type': infer_pack_uom_type(uom_label, unit_packaging),
        })
    else:
        vals['pack_qty'] = 0.0

    if categ_name:
        vals['_rdl_category_name'] = categ_name
    return vals


def import_products(env, file_bytes, *, sheet_index=SHEET_PRODUCT_MASTER, header_row=DEFAULT_HEADER_ROW_1INDEXED):
    """Import/update products from the Product Master sheet."""
    workbook = load_workbook(file_bytes)
    sheet = resolve_sheet(workbook, sheet_index=sheet_index)
    Product = env['product.template']
    Category = env['product.category']

    created = updated = skipped = 0
    for _line_no, row in iter_data_rows(sheet, header_row):
        product_vals = build_product_vals(row)
        if not product_vals:
            skipped += 1
            continue

        categ_name = product_vals.pop('_rdl_category_name', None)
        if categ_name:
            category = Category.search([('name', '=ilike', categ_name)], limit=1)
            if not category:
                category = Category.create({'name': categ_name})
            product_vals['categ_id'] = category.id

        existing = Product.search([('default_code', '=', product_vals['default_code'])], limit=1)
        if existing:
            existing.write(product_vals)
            updated += 1
        else:
            Product.create(product_vals)
            created += 1

    return {'created': created, 'updated': updated, 'skipped': skipped}


def resolve_stock_location(env, warehouse_hint: str):
    """Resolve internal stock location from warehouse/location hint or default WH."""
    company = env.company
    warehouse_hint = safe_str(warehouse_hint)
    Warehouse = env['stock.warehouse']
    Location = env['stock.location']

    if warehouse_hint:
        if '/' in warehouse_hint:
            loc = Location.search([
                ('complete_name', 'ilike', warehouse_hint),
                ('usage', '=', 'internal'),
                '|', ('company_id', '=', company.id), ('company_id', '=', False),
            ], limit=1)
            if loc:
                return loc

        wh = Warehouse.search([
            '|', ('code', '=ilike', warehouse_hint), ('name', 'ilike', warehouse_hint),
            ('company_id', '=', company.id),
        ], limit=1)
        if wh and wh.lot_stock_id:
            return wh.lot_stock_id

    wh = Warehouse.search([('company_id', '=', company.id)], limit=1)
    if not wh or not wh.lot_stock_id:
        from odoo.exceptions import UserError
        raise UserError(("Could not find a default warehouse/stock location for %s.") % company.name)
    return wh.lot_stock_id


def apply_inventory_quantity(env, product, location, qty, unit_cost=None):
    """Set on-hand quantity via standard inventory adjustment."""
    if unit_cost and unit_cost > 0:
        product.product_tmpl_id.with_context(disable_auto_svl=True).write({
            'standard_price': unit_cost,
        })

    Quant = env['stock.quant'].sudo().with_context(inventory_mode=True)
    quant = Quant.search([
        ('product_id', '=', product.id),
        ('location_id', '=', location.id),
    ], limit=1)
    if not quant:
        quant = Quant.create({
            'product_id': product.id,
            'location_id': location.id,
        })
    quant.inventory_quantity = qty
    quant.action_apply_inventory()


def import_opening_inventory(
    env,
    file_bytes,
    *,
    sheet_index=SHEET_OPENING_INVENTORY,
    header_row=DEFAULT_HEADER_ROW_1INDEXED,
):
    """Load opening stock from the Opening Inventory sheet."""
    workbook = load_workbook(file_bytes)
    sheet = resolve_sheet(workbook, sheet_index=sheet_index)
    Product = env['product.product']

    applied = skipped = missing = 0
    aggregates: dict[tuple[int, int], dict[str, float]] = {}

    for _line_no, row in iter_data_rows(sheet, header_row):
        sku = safe_str(pick(row, 'SKU'))
        if not sku:
            skipped += 1
            continue

        product = Product.search([('default_code', '=', sku)], limit=1)
        if not product:
            _logger.warning("Opening inventory: SKU %s not found — import products first.", sku)
            missing += 1
            continue

        qty = pick_float(row, 'Quantity On Hand', 'Qty On Hand', 'Quantity')
        if qty <= 0:
            skipped += 1
            continue

        unit_cost = pick_float(row, 'Unit Cost (₦)', 'Unit Cost', 'Cost Price (₦)', 'Cost Price')
        warehouse_hint = safe_str(pick(row, 'Warehouse', 'Location'))
        location = resolve_stock_location(env, warehouse_hint)
        key = (product.id, location.id)
        bucket = aggregates.setdefault(key, {'qty': 0.0, 'unit_cost': 0.0, 'rows': 0})
        bucket['qty'] += qty
        if unit_cost:
            bucket['unit_cost'] += unit_cost
            bucket['rows'] += 1

    for (product_id, location_id), data in aggregates.items():
        product = Product.browse(product_id)
        location = env['stock.location'].browse(location_id)
        unit_cost = data['unit_cost'] / data['rows'] if data['rows'] else 0.0
        apply_inventory_quantity(env, product, location, data['qty'], unit_cost or None)
        applied += 1

    return {'applied': applied, 'skipped': skipped, 'missing': missing}
