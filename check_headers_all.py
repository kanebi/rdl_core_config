import openpyxl
import json

wb = openpyxl.load_workbook('RDL_Trading_Odoo.xlsx', data_only=True)
for i, sheet_name in enumerate(wb.sheetnames):
    print(f'\n--- Sheet {i}: {sheet_name} ---')
    sheet = wb[sheet_name]
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx > 5: break
        print(f'Row {r_idx}:', [str(c).strip() if c else '' for c in row][:15])
