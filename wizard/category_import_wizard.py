# -*- coding: utf-8 -*-
import base64
import logging
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import io

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None

class CategoryImportWizard(models.TransientModel):
    _name = 'category.import.wizard'
    _description = 'Import Product Categories from Excel'

    excel_file = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='File Name')
    sheet_index = fields.Integer(string='Sheet Index', default=5, required=True, help="0-based index of the sheet to read.")
    header_row_index = fields.Integer(string='Header Row Index', default=3, required=True, help="0-based index of the header row.")

    def action_import_categories(self):
        if not openpyxl:
            raise UserError(_("Please install the 'openpyxl' Python library to import Excel files."))

        try:
            file_data = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise UserError(_('Error reading the Excel file: %s') % str(e))

        sheet_names = wb.sheetnames
        if self.sheet_index < 0 or self.sheet_index >= len(sheet_names):
            raise UserError(_('Invalid Sheet Index. The file only has %d sheets.') % len(sheet_names))

        sheet = wb[sheet_names[self.sheet_index]]
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) <= self.header_row_index:
            raise UserError(_('The specified Header Row Index is out of range.'))

        header_row = rows[self.header_row_index]
        name_col = None
        parent_col = None

        for col_idx, col_name in enumerate(header_row):
            if not col_name: continue
            col_name_str = str(col_name).strip().lower()
            if col_name_str == 'category name':
                name_col = col_idx
            elif col_name_str == 'parent category':
                parent_col = col_idx

        if name_col is None:
            raise UserError(_('Could not find required column "Category Name" in the header row.'))

        ProductCategory = self.env['product.category']
        PosCategory = self.env['pos.category']
        ProductTemplate = self.env['product.template']

        # Helper to recursively get or create product.category
        def get_or_create_product_categ(cat_name, parent_name=None):
            if not cat_name: return False
            cat_name = str(cat_name).strip()
            
            # Check if exists
            existing = ProductCategory.search([('name', '=ilike', cat_name)], limit=1)
            if existing:
                # Update parent if provided and different
                if parent_name and not existing.parent_id:
                    parent_categ = get_or_create_product_categ(parent_name)
                    if parent_categ and parent_categ.id != existing.id:
                        existing.parent_id = parent_categ.id
                return existing
            
            # Create new
            parent_id = False
            if parent_name:
                parent_categ = get_or_create_product_categ(parent_name)
                if parent_categ:
                    parent_id = parent_categ.id
                    
            return ProductCategory.create({
                'name': cat_name,
                'parent_id': parent_id
            })

        # Helper to recursively get or create pos.category
        def get_or_create_pos_categ(cat_name, parent_id=False):
            if not cat_name: return False
            cat_name = str(cat_name).strip()
            
            existing = PosCategory.search([('name', '=ilike', cat_name)], limit=1)
            if existing:
                if parent_id and not existing.parent_id:
                    existing.parent_id = parent_id
                return existing
                
            return PosCategory.create({
                'name': cat_name,
                'parent_id': parent_id
            })

        success_count = 0
        
        # Read Excel rows and build the hierarchy
        for r_idx in range(self.header_row_index + 1, len(rows)):
            row = rows[r_idx]
            cat_name = row[name_col] if name_col < len(row) else None
            parent_name = row[parent_col] if parent_col is not None and parent_col < len(row) else None

            if not cat_name or not str(cat_name).strip():
                continue

            # 1. Create/Update Product Category
            prod_categ = get_or_create_product_categ(cat_name, parent_name)
            
            # 2. Create/Update POS Category (Mirror hierarchy)
            pos_parent_id = False
            if prod_categ.parent_id:
                pos_parent = get_or_create_pos_categ(prod_categ.parent_id.name)
                pos_parent_id = pos_parent.id if pos_parent else False
                
            pos_categ = get_or_create_pos_categ(cat_name, pos_parent_id)

            # 3. Assign pos.category to corresponding products
            if prod_categ and pos_categ:
                products = ProductTemplate.search([('categ_id', '=', prod_categ.id)])
                for prod in products:
                    if pos_categ.id not in prod.pos_categ_ids.ids:
                        prod.write({'pos_categ_ids': [(4, pos_categ.id)]})

            success_count += 1

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Categories Imported'),
                'message': _('%d categories have been imported and synced to POS.') % success_count,
                'sticky': False,
                'type': 'success',
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }
