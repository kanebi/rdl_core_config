# -*- coding: utf-8 -*-
import base64
import logging
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import io

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None

ACCOUNT_TYPE_MAP = {
    'current assets': 'asset_current',
    'bank and cash': 'asset_cash',
    'cash': 'asset_cash',
    'non-current assets': 'asset_non_current',
    'fixed assets': 'asset_fixed',
    'current liabilities': 'liability_current',
    'non-current liabilities': 'liability_non_current',
    'equity': 'equity',
    'income': 'income',
    'other income': 'income_other',
    'expense': 'expense',
    'expenses': 'expense',
    'cost of revenue': 'expense_direct_cost',
    'depreciation': 'expense_depreciation',
}

class AccountImportWizard(models.TransientModel):
    _name = 'account.import.wizard'
    _description = 'Import Chart of Accounts from Excel'

    excel_file = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='File Name')
    sheet_index = fields.Integer(string='Sheet Index', default=6, required=True, help="0-based index of the sheet to read.")
    header_row_index = fields.Integer(string='Header Row Index', default=3, required=True, help="0-based index of the header row.")

    def action_import_accounts(self):
        if not openpyxl:
            raise UserError(_("Please install the 'openpyxl' Python library to import Excel files."))

        try:
            file_data = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise UserError(_('Error reading the Excel file: %s') % str(e))

        sheet_names = wb.sheetnames
        if self.sheet_index < 0 or self.sheet_index >= len(sheet_names):
            raise UserError(_('Invalid Sheet Index. The file only has %d sheets.') % len(sheet_names))

        sheet = wb[sheet_names[self.sheet_index]]
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) <= self.header_row_index:
            raise UserError(_('The specified Header Row Index is out of range.'))

        header_row = rows[self.header_row_index]
        code_col = None
        name_col = None
        type_col = None
        parent_col = None

        for col_idx, col_name in enumerate(header_row):
            if not col_name: continue
            col_name_str = str(col_name).strip().lower()
            if col_name_str == 'account code':
                code_col = col_idx
            elif col_name_str == 'account name':
                name_col = col_idx
            elif col_name_str == 'account type':
                type_col = col_idx
            elif col_name_str == 'parent account':
                parent_col = col_idx

        if code_col is None or name_col is None:
            raise UserError(_('Could not find required columns "Account Code" or "Account Name" in the header row.'))

        Account = self.env['account.account']
        AccountGroup = self.env['account.group']
        Company = self.env.company

        # Helper to get or create Account Group
        def get_or_create_group(group_code):
            if not group_code: return False
            group_code = str(group_code).strip()
            group = AccountGroup.search([('code_prefix_start', '<=', group_code), ('code_prefix_end', '>=', group_code)], limit=1)
            if not group:
                group = AccountGroup.search([('name', '=', group_code)], limit=1)
            if not group:
                group = AccountGroup.create({
                    'name': f"Group {group_code}",
                    'code_prefix_start': group_code,
                    'code_prefix_end': group_code,
                })
            return group

        success_count = 0
        has_parent_id_field = hasattr(Account, 'parent_id')

        # Read Excel rows and build the accounts
        for r_idx in range(self.header_row_index + 1, len(rows)):
            row = rows[r_idx]
            code = row[code_col] if code_col < len(row) else None
            name = row[name_col] if name_col < len(row) else None
            acc_type_raw = row[type_col] if type_col is not None and type_col < len(row) else None
            parent = row[parent_col] if parent_col is not None and parent_col < len(row) else None

            if not code or not str(code).strip():
                continue

            code_str = str(code).strip()
            name_str = str(name).strip() if name else ""
            acc_type_str = str(acc_type_raw).strip().lower() if acc_type_raw else ""
            
            # Map type or default to asset_current
            mapped_type = ACCOUNT_TYPE_MAP.get(acc_type_str, 'asset_current')

            parent_str = str(parent).strip() if parent else False

            # Find existing account
            account = Account.search([('code', '=', code_str), ('company_ids', 'in', [Company.id])], limit=1)
            
            vals = {
                'name': name_str,
                'account_type': mapped_type,
            }

            if parent_str:
                if has_parent_id_field:
                    parent_acc = Account.search([('code', '=', parent_str), ('company_ids', 'in', [Company.id])], limit=1)
                    if parent_acc:
                        vals['parent_id'] = parent_acc.id
                else:
                    # Fallback to group_id
                    group = get_or_create_group(parent_str)
                    if group:
                        vals['group_id'] = group.id

            if account:
                account.write(vals)
            else:
                vals['code'] = code_str
                vals['company_ids'] = [(4, Company.id)]
                Account.create(vals)

            success_count += 1

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Accounts Imported'),
                'message': _('%d accounts have been imported.') % success_count,
                'sticky': False,
                'type': 'success',
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }
