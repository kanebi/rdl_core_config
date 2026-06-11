# -*- coding: utf-8 -*-
import base64
import logging
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import io

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None

class InventoryImportWizard(models.TransientModel):
    _name = 'inventory.import.wizard'
    _description = 'Import Opening Inventory from Excel'

    excel_file = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='File Name')
    sheet_index = fields.Integer(string='Sheet Index', default=3, required=True, help="0-based index of the sheet to read.")
    header_row_index = fields.Integer(string='Header Row Index', default=3, required=True, help="0-based index of the header row.")

    def action_import_inventory(self):
        if not openpyxl:
            raise UserError(_("Please install the 'openpyxl' Python library to import Excel files."))

        try:
            file_data = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise UserError(_('Error reading the Excel file: %s') % str(e))

        sheet_names = wb.sheetnames
        if self.sheet_index < 0 or self.sheet_index >= len(sheet_names):
            raise UserError(_('Invalid Sheet Index. The file only has %d sheets.') % len(sheet_names))

        sheet = wb[sheet_names[self.sheet_index]]
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) <= self.header_row_index:
            raise UserError(_('The specified Header Row Index is out of range.'))

        header_row = rows[self.header_row_index]
        sku_col = None
        qty_col = None

        for col_idx, col_name in enumerate(header_row):
            if not col_name: continue
            col_name_str = str(col_name).strip().lower()
            if col_name_str == 'sku':
                sku_col = col_idx
            elif col_name_str == 'quantity on hand':
                qty_col = col_idx

        if sku_col is None or qty_col is None:
            raise UserError(_('Could not find required columns "SKU" or "Quantity On Hand" in the header row.'))

        main_wh = self.env['stock.warehouse'].search([('company_id', '=', self.env.company.id)], limit=1)
        if not main_wh:
            raise UserError(_('Could not find a default warehouse for the current company.'))
        default_location_id = main_wh.lot_stock_id.id

        Quant = self.env['stock.quant']
        Product = self.env['product.product']

        success_count = 0
        quantities_to_apply = {} # {product_id: qty}
        
        for r_idx in range(self.header_row_index + 1, len(rows)):
            row = rows[r_idx]
            sku = row[sku_col] if sku_col < len(row) else None
            qty_val = row[qty_col] if qty_col < len(row) else None

            if not sku or not str(sku).strip():
                continue

            sku_str = str(sku).strip()
            product = Product.search([('default_code', '=', sku_str)], limit=1)

            if not product:
                _logger.warning("SKU %s not found in products. Skipping.", sku_str)
                continue

            try:
                qty = float(qty_val) if qty_val else 0.0
            except ValueError:
                qty = 0.0

            if qty == 0:
                continue

            if product.is_brewery:
                if product.liquid_product_id:
                    liquid_qty = qty * (product.brewery_liquid_qty or 1.0)
                    quantities_to_apply[product.liquid_product_id.id] = quantities_to_apply.get(product.liquid_product_id.id, 0.0) + liquid_qty
                if product.bottle_product_id:
                    bottle_qty = qty * (product.brewery_bottle_qty or 1.0)
                    quantities_to_apply[product.bottle_product_id.id] = quantities_to_apply.get(product.bottle_product_id.id, 0.0) + bottle_qty
                if product.crate_product_id:
                    crate_qty = qty * getattr(product, 'brewery_crate_qty', 1.0)
                    quantities_to_apply[product.crate_product_id.id] = quantities_to_apply.get(product.crate_product_id.id, 0.0) + crate_qty
            else:
                quantities_to_apply[product.id] = quantities_to_apply.get(product.id, 0.0) + qty

        for prod_id, total_qty in quantities_to_apply.items():
            if total_qty <= 0: continue

            # Create or update stock quant (Inventory Adjustment)
            quant = Quant.search([
                ('product_id', '=', prod_id),
                ('location_id', '=', default_location_id)
            ], limit=1)

            if not quant:
                quant = Quant.create({
                    'product_id': prod_id,
                    'location_id': default_location_id,
                    'inventory_quantity': total_qty,
                })
            else:
                quant.inventory_quantity = total_qty

            quant.action_apply_inventory()
            success_count += 1

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Successful'),
                'message': _('%d inventory lines have been imported and applied.') % success_count,
                'sticky': False,
                'type': 'success',
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }
