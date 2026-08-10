# -*- coding: utf-8 -*-
import base64
import io
from odoo import fields, models, _
from odoo.exceptions import UserError
import openpyxl


class ProductImportWizard(models.TransientModel):
    _name = 'product.import.wizard'
    _description = 'Product Import Wizard'

    file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='File Name')

    sheet_name = fields.Char(
        string='Sheet Name or Index',
        default="1",
        help="Enter sheet name or 0-based index (e.g. 1 for second sheet)",
    )
    start_row = fields.Integer(
        string='Header Row',
        default=4,
        help="Row number where the headers are located (1-indexed, e.g., 4)",
    )

    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Please upload an Excel file."))

        try:
            file_content = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

            sheet_identifier = self.sheet_name.strip() if self.sheet_name else '0'
            if sheet_identifier.isdigit():
                sheet_idx = int(sheet_identifier)
                if sheet_idx < len(wb.worksheets):
                    sheet = wb.worksheets[sheet_idx]
                else:
                    raise UserError(_("Sheet index out of bounds."))
            else:
                if sheet_identifier in wb.sheetnames:
                    sheet = wb[sheet_identifier]
                else:
                    raise UserError(_("Sheet name '%s' not found in workbook.") % sheet_identifier)

        except UserError:
            raise
        except Exception as e:
            raise UserError(_("Error reading the Excel file: %s") % str(e))

        Product = self.env['product.template']
        Category = self.env['product.category']

        rows = list(sheet.iter_rows(values_only=True))
        header_row_idx = self.start_row - 1
        if header_row_idx < 0 or header_row_idx >= len(rows):
            raise UserError(_("Header row is out of bounds."))

        headers = [str(h).strip() if h else '' for h in rows[header_row_idx]]

        def safe_float(val):
            if val is None or str(val).strip() == '' or str(val).lower() == 'nan':
                return 0.0
            try:
                return float(val)
            except ValueError:
                return 0.0

        for r_idx in range(header_row_idx + 1, len(rows)):
            row_data = rows[r_idx]
            row = dict(zip(headers, row_data))

            sku_val = row.get('SKU')
            if sku_val is None or str(sku_val).strip() == '' or str(sku_val).lower() == 'nan':
                continue

            sku = str(sku_val).strip()
            name = str(row.get('Product Name', '')).strip()
            unit_packaging = str(row.get('Unit Packaging', '')).strip()
            categ_name = str(row.get('Category', '')).strip()

            categ = False
            if categ_name and str(categ_name).lower() != 'nan':
                categ = Category.search([('name', '=ilike', categ_name)], limit=1)
                if not categ:
                    categ = Category.create({'name': categ_name})

            existing_product = Product.search([('default_code', '=', sku)], limit=1)

            product_vals = {
                'name': name,
                'default_code': sku,
                'type': 'consu',
                'is_storable': True,
            }
            if categ:
                product_vals['categ_id'] = categ.id

            qty_in_pack = safe_float(row.get('Bottles in a Crate'))
            sales_price = safe_float(row.get('Sales Price (₦)'))
            cost_price = safe_float(row.get('Cost Price (₦)'))
            if not cost_price:
                cost_price = safe_float(row.get('Full Crate Cost Price'))

            if unit_packaging.lower() == 'bottle':
                # One finished SKU. Excel prices are for the full pack; product UoM is Crate xN.
                pack_qty = qty_in_pack or 24.0
                product_vals.update({
                    'is_brewery': True,
                    'is_packaged_drinks': False,
                    'pack_qty': pack_qty,
                    'pack_uom_type': 'crate',
                    'list_price': sales_price,
                    'standard_price': cost_price,
                })
            elif unit_packaging.lower() in ['can', 'plastic', 'pet', 'carton']:
                pack_qty = qty_in_pack or 24.0
                product_vals.update({
                    'is_brewery': False,
                    'is_packaged_drinks': True,
                    'pack_qty': pack_qty,
                    'pack_uom_type': 'carton',
                    'list_price': sales_price,
                    'standard_price': cost_price,
                })
            else:
                if sales_price or cost_price:
                    product_vals.update({
                        'list_price': sales_price,
                        'standard_price': cost_price,
                    })

            if existing_product:
                existing_product.write(product_vals)
            else:
                Product.create(product_vals)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Successful'),
                'message': _('Products have been imported successfully.'),
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }
